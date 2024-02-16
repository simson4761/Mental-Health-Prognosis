# This is a sample project

import csv
import os
import platform
import sqlite3
import pandas as pd
import nltk
import openpyxl
import pyphen as pyphen
from newspaper import Article
from nltk.corpus import opinion_lexicon
from nltk.corpus import stopwords
from nltk.tokenize import RegexpTokenizer
from nltk.tokenize import sent_tokenize
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.model_selection import train_test_split
from sklearn.model_selection import RepeatedStratifiedKFold
from sklearn.model_selection import cross_val_score
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn import datasets
import matplotlib.pyplot as plt

# Get the operating system information
os_name = platform.system()

# Get the current username from the USER environment variable
username = os.getenv('USER')

# Path to the Chrome history database

# For Windows ↓
history_db_path_windows = fr"C:\Users\{username}\AppData\Local\Google\Chrome\User Data\Default\History"

# For Mac ↓
history_db_path_mac = f"/Users/{username}/Library/Application Support/Google/Chrome/Default/History"

# Connect to the Chrome history database
if os_name == "Darwin":
    connection = sqlite3.connect(history_db_path_mac)
else:
    connection = sqlite3.connect(history_db_path_windows)

cursor = connection.cursor()

# Query to retrieve browser history
query = "SELECT title, url, visit_count,  last_visit_time FROM urls ORDER BY last_visit_time DESC"

# Execute the query
cursor.execute(query)

# Fetch all the rows
history_data = cursor.fetchall()

# Close the database connection
connection.close()

# Write the history data to a CSV file
csv_browser_history = "chrome_history.csv"
with open(csv_browser_history, "w", newline="", encoding="utf-8") as csv_file:
    csv_writer = csv.writer(csv_file)
    csv_writer.writerow(["Title", "URL", "Visit Count", "Last Visit Time"])
    for row in history_data:
        csv_writer.writerow(row)

# print("Browser history extracted and saved to", csv_browser_history)

# xlsx = openpyxl.load_workbook(csv_browser_history)  # Insert the input file location
# sheet = xlsx.active

output_sheet = openpyxl.Workbook()
op_sheet = output_sheet.active
op_sheet.title = "outputSheet"


def filterStopWords(text):
    stop_words = set(stopwords.words('english'))
    word_tokens = nltk.word_tokenize(text)
    filtered_sentence = []
    for w in word_tokens:
        if w not in stop_words:
            filtered_sentence.append(w)
    return filtered_sentence


def positive_score_of(filtered_text):
    positive_words = set(opinion_lexicon.positive())
    positive_score = 0
    for w in filtered_text:
        if w.lower() in positive_words:
            positive_score = positive_score + 1
    return positive_score


def negative_score_of(filtered_text):
    negative_words = set(opinion_lexicon.negative())
    negative_score = 0
    for w in filtered_text:
        if w.lower() in negative_words:
            negative_score = negative_score + 1
    return negative_score


def polarity_of(positive_score, negative_score):
    return (positive_score - negative_score) / (positive_score + negative_score + 0.000001)


def subjectivity_of(positive_score, negative_score, length):
    return (positive_score + negative_score) / (length + 0.000001)


def average_sentence_length_of(text):
    global i
    sent_tokenize_list = sent_tokenize(text)
    word_count = 0
    for i in range(0, len(sent_tokenize_list)):
        word_count += sent_tokenize_list[i].count(" ")
    if i > 0:
        return word_count / i
    else:
        return word_count / (i + 1)


def average_no_of_words_per_sentence_length_of(text):
    word_tokenize = nltk.word_tokenize(text)
    sent_tokenize_list = sent_tokenize(text)
    if len(sent_tokenize_list) > 0:
        return len(word_tokenize) / len(sent_tokenize_list)
    else:
        return len(word_tokenize) / (len(sent_tokenize_list) + 1)


def word_count_of(text):
    tokenizer = RegexpTokenizer(r'\w+')
    return len(tokenizer.tokenize(text))


def complex_word_count_of(text, word_count):
    tokenizer = RegexpTokenizer(r'\w+')
    complex_word_count = 0
    total_syllable = 0
    word_tokenizer = tokenizer.tokenize(text)
    for ele in range(0, len(word_tokenizer)):
        dic = pyphen.Pyphen(lang='en')
        count1 = dic.inserted(word_tokenizer[ele]).count("-") + 1
        total_syllable = total_syllable + dic.inserted(word_tokenizer[ele]).count("-") + 1
        if count1 > 2:
            complex_word_count = complex_word_count + 1
    syllable_per_word = total_syllable / max(word_count, 1)
    return complex_word_count, syllable_per_word


def personal_pronoun_of(text):
    word_tokens = nltk.word_tokenize(text)
    tagged = nltk.pos_tag(word_tokens)
    prp_tag = [t for t in tagged if t[1] == "PRP"]
    return len(prp_tag)


def average_word_length_of(text, word_count):
    tokenizer = RegexpTokenizer(r'\w+')
    word_tokens = tokenizer.tokenize(text)
    character_length = 0
    for m in range(0, len(word_tokens)):
        character_length = character_length + len(word_tokens[m])
    if word_count > 0:
        return character_length / word_count
    else:
        return character_length / (word_count + 1)


def sentimentAnalysis(text, output_cell_address, urlsample):
    filtered_text = filterStopWords(text)
    positive_score = positive_score_of(filtered_text)
    negative_score = negative_score_of(filtered_text)
    polarity_score = polarity_of(positive_score, negative_score)
    subjectivity_score = subjectivity_of(positive_score, negative_score, len(filtered_text))
    average_sentence_length = average_sentence_length_of(text)
    word_count = word_count_of(text)
    complex_word_count, syllable_per_word = complex_word_count_of(text, word_count)
    percentage_of_complex_words = complex_word_count / max(word_count, 1)
    average_no_of_words_per_sentence_length = average_no_of_words_per_sentence_length_of(text)
    fog_index = (average_sentence_length + percentage_of_complex_words) * 0.4
    personal_pronoun_count = personal_pronoun_of(text)
    average_word_length = average_word_length_of(text, word_count)
    # print("positive score", average_word_length)
    if average_word_length > 0:
        op_sheet.cell(output_cell_address, 1).value = output_cell_address
        op_sheet.cell(output_cell_address, 2).value = urlsample
        op_sheet.cell(output_cell_address, 3).value = positive_score
        op_sheet.cell(output_cell_address, 4).value = negative_score
        op_sheet.cell(output_cell_address, 5).value = polarity_score
        op_sheet.cell(output_cell_address, 6).value = subjectivity_score
        op_sheet.cell(output_cell_address, 7).value = average_sentence_length
        op_sheet.cell(output_cell_address, 8).value = percentage_of_complex_words
        op_sheet.cell(output_cell_address, 9).value = fog_index
        op_sheet.cell(output_cell_address, 10).value = average_no_of_words_per_sentence_length
        op_sheet.cell(output_cell_address, 11).value = complex_word_count
        op_sheet.cell(output_cell_address, 12).value = word_count
        op_sheet.cell(output_cell_address, 13).value = syllable_per_word
        op_sheet.cell(output_cell_address, 14).value = personal_pronoun_count
        op_sheet.cell(output_cell_address, 15).value = average_word_length
    else:
        op_sheet.cell(output_cell_address, 1).value = "N/A"
        op_sheet.cell(output_cell_address, 2).value = "N/A"
        op_sheet.cell(output_cell_address, 3).value = "N/A"
        op_sheet.cell(output_cell_address, 4).value = "N/A"
        op_sheet.cell(output_cell_address, 5).value = "N/A"
        op_sheet.cell(output_cell_address, 6).value = "N/A"
        op_sheet.cell(output_cell_address, 7).value = "N/A"
        op_sheet.cell(output_cell_address, 8).value = "N/A"
        op_sheet.cell(output_cell_address, 9).value = "N/A"
        op_sheet.cell(output_cell_address, 10).value = "N/A"
        op_sheet.cell(output_cell_address, 11).value = "N/A"
        op_sheet.cell(output_cell_address, 12).value = "N/A"
        op_sheet.cell(output_cell_address, 13).value = "N/A"
        op_sheet.cell(output_cell_address, 14).value = "N/A"
        op_sheet.cell(output_cell_address, 15).value = "N/A"


def output_formatting():
    op_sheet.cell(1, 1).value = "URL_ID"
    op_sheet.cell(1, 2).value = "URL"
    op_sheet.cell(1, 3).value = "POSITIVE SCORE"
    op_sheet.cell(1, 4).value = "NEGATIVE SCORE"
    op_sheet.cell(1, 5).value = "POLARITY SCORE"
    op_sheet.cell(1, 6).value = "SUBJECTIVITY SCORE"
    op_sheet.cell(1, 7).value = "AVG SENTENCE LENGTH"
    op_sheet.cell(1, 8).value = "PERCENTAGE OF COMPLEX WORDS"
    op_sheet.cell(1, 9).value = "FOG INDEX"
    op_sheet.cell(1, 10).value = "AVG NUMBER OF WORDS PER SENTENCE"
    op_sheet.cell(1, 11).value = "COMPLEX WORD COUNT"
    op_sheet.cell(1, 12).value = "WORD COUNT"
    op_sheet.cell(1, 13).value = "SYLLABLE PER WORD"
    op_sheet.cell(1, 14).value = "PERSONAL PRONOUNS"
    op_sheet.cell(1, 15).value = "AVG WORD LENGTH"


def articleCleanup(urlSample, cell_address):
    try:

        article = Article("" + urlSample + "")
        article.download()
        article.parse()
        article.nlp()

        # Find the main article content based on HTML tags, classes, or IDs specific to the website
        # You may need to customize this part based on the structure of the website you're working with
        article_content = article.text
        if article_content:
            # Extract and show the text of the article
            text = article_content
            output_formatting()
            if len(text) > 0:
                sentimentAnalysis(text, cell_address, urlSample)
            else:
                print("EMPTY")
            print(cell_address)
        else:
            print("Article content not found on the page. You may need to customize the parsing logic.")

    except Exception as e:
        print(e)
        # op_sheet.cell(cell_address, 1).value = cell_address
        # op_sheet.cell(cell_address, 2).value = urlSample
        # for op_i in range(3, 16):
        #     op_sheet.cell(cell_address, op_i).value = "404 Error"


# for i in range(2, 'chrome_history.csv'.max_row + 1):
#     output_formatting()

#     url = csv_browser_history.cell(i, 2).value
#     op_sheet.cell(i, 1).value = csv_browser_history.cell(i, 1).value
#     op_sheet.cell(i, 2).value = url

with open('chrome_history.csv', 'r', newline='') as csv_file:
    # Create a CSV reader
    csv_reader = csv.reader(csv_file)

    k = 1
    # Read and process each row
    for row in csv_reader:
        # Access individual columns by index
        title, url, count, time = row
        articleCleanup(url, k)
        k += 1


def checkEmpty(outSheet, n):
    if outSheet.cell(row=n, column=1).value is None:
        outSheet.delete_rows(n, 1)


for j in range(1, op_sheet.max_row + 1):
    for k in range(1, op_sheet.max_row + 1):
        checkEmpty(op_sheet, k)

# print(len(op_sheet.index()))
output_sheet.save("/Users/simson/Downloads/output_sheet_assignment.csv")
x = []
y = []
for i in range(2, op_sheet.max_row + 1):
    x1 = []
    y1 = []

    x1.append(float(op_sheet.cell(i, 3).value))
    x1.append(float(op_sheet.cell(i, 4).value))
    x1.append(float(op_sheet.cell(i, 5).value))
    x1.append(float(op_sheet.cell(i, 6).value))
    x1.append(float(op_sheet.cell(i, 7).value))
    x1.append(float(op_sheet.cell(i, 10).value))
    x1.append(float(op_sheet.cell(i, 12).value))

    x.append(x1)
    y.append(float(op_sheet.cell(i, 9).value))

lda = LinearDiscriminantAnalysis()
print(x, y)
# Fit the model
lda.fit(x, y)
iris = datasets.load_iris()
data_plot = lda.fit(x, y).transform(y)
target_names = iris.target_names

# create LDA plot
plt.figure()
colors = ['red', 'green', 'blue']
lw = 2
for color, i, target_name in zip(colors, [0, 1, 2], target_names):
    plt.scatter(data_plot[y == i, 0], data_plot[y == i, 1], alpha=.8, color=color,
                label=target_name)

# add legend to plot
plt.legend(loc='best', shadow=False, scatterpoints=1)

# display LDA plot
plt.show()
